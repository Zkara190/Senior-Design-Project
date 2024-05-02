import math
import openpyxl
import time
from selenium import webdriver

test_vals = [0,0,0,0,0,0]
output_s = [10,20,30,40,50,60]
ids = ["b_dot","y_dot","t_dot"]
id_correct = ["t_dot","y_dot","b_dot"]

def align_vals(ids, id_correct, output_s):
    #int to increment for the loop
    l = 0

    for k in range (0,3):
        #loop check for the correct order of the ids and assign the out floats to the correct index of test vals
        if ids [0] == id_correct[k]:
            test_vals [l] = output_s [0]
            test_vals [l+1] = output_s [1]
        elif ids [1] == id_correct[k]:
            test_vals [l] = output_s [2]
            test_vals [l+1] = output_s [3]
        else :  
            test_vals [l] = output_s [4]
            test_vals [l+1] = output_s [5]
        l += 2




align_vals(ids, id_correct, output_s)
print("unaligned: ", output_s)
print("order of ids: ", ids)
print("correct: ", id_correct)
print("aligned: ", test_vals)

x = 0
y=2
#load excel file
wb = openpyxl.load_workbook(r'C:\Users\awind\Desktop\test1.xlsx')
sheet = wb.active


for x in range(0,1):
    #create new sheet within excel




    #random assigned values for test
    x_b = test_vals[x]
    y_b = test_vals[x+1]
    x_r = test_vals[x+2]
    y_r = test_vals[x+3]
    x_target = test_vals[x+4]
    y_target = test_vals[x+5]
  



    #calculate the angle the car is pointing based on red and blue dots
    #blue is front
    def car_angle(x_b,y_b,x_r,y_r):
        return math.degrees(math.atan((y_b-y_r)/(x_b-x_r)))


    #Calculate the distance from the blue part of VATR to the target
    def car_distance(x_b,y_b,x_target,y_target):
        return math.sqrt((x_target-x_b)**2+(y_target-y_b)**2)


    #angle from car blue side to the target
    def car_to_target_angle(x_b,y_b,x_target,y_target):
        return math.degrees(math.atan((y_target-y_b)/(x_target-x_b)))


    #calculate the desired angle to turn for the VATR
    def desired_angle(car_to_target_angle,car_angle):
           return (car_to_target_angle(x_b,y_b,x_target,y_target)-car_angle(x_b,y_b,x_r,y_r))




    #temp values to assist in calculations
    tmp = desired_angle(car_to_target_angle,car_angle)
    tmp_dist = car_distance(x_b,y_b,x_target,y_target)

    #assign true or false if the desired angle is within range or not
    #true means drive straight
    #false means turn
    if -5.0 < tmp < 5.0:
        forward_over_turn = True
    else:
        forward_over_turn = False




    #set output value if VATR needs to drive straight
    def straight(car_distance):
        if -5 >= tmp_dist >= -10:
            return 16
        elif -11 >= tmp_dist >= -20:
            return 15
        elif -21 >= tmp_dist >= -30:
            return 14
        elif 5 <= tmp_dist <= 10:
            return 16
        elif 11 <= tmp_dist <= 20:
            return 15
        elif 21 <= tmp_dist <= 30:
            return 14
        else: 
            return 7


    #neg is left
    #pos is right
    #-5 to 5 is stop
    #set output value if VATR needs to turn
    def switch(tmp):
        if -5.0 >= tmp >= -65.0:
            return 6
        elif -66.0 >= tmp >= -125.0:
            return 4
        elif -126.0 >= tmp >= -185.0:
            return 2
        elif -186.0 >= tmp >= -245.0:
            return 8
        elif -246.0 >= tmp >= -305.0:
            return 10
        elif -306.0 >= tmp >= -360.0:
            return 12
        elif 5.0 <= tmp <= 65.0:
            return 5
        elif 66.0 <= tmp <= 125.0:
            return 3
        elif 126.0 <= tmp <= 185.0:
            return 1
        elif 186.0 <= tmp <= 245.0:
            return 9
        elif 246.0 <= tmp <= 305.0:
            return 11
        elif 306.0 <= tmp <= 360.0:
            return 13
        elif -5 < tmp < 5:
            return 7
        else:
            return 20

    #assign out value so the VATR needs to turn or drive straight
    def output(forward_over_turn):
        if forward_over_turn == False:
            return switch(tmp)
        elif forward_over_turn == True:
            return straight(car_distance)

    #final descision value command
    out = output(forward_over_turn)


    #ouput values of steps
    print("Car Angle = ", car_angle(x_b,y_b,x_r,y_r))
    print("Car Distance = ", car_distance(x_b,y_b,x_target,y_target), " m")
    print("Car Distance = ", car_to_target_angle(x_b,y_b,x_target,y_target))
    print("Car Distance = ", desired_angle(car_to_target_angle,car_angle))
    print("Output = ", out)


    data_values = (x+1),car_angle(x_b,y_b,x_r,y_r),car_distance(x_b,y_b,x_target,y_target),car_to_target_angle(x_b,y_b,x_target,y_target),desired_angle(car_to_target_angle,car_angle),x_b,y_b,x_r,y_r,x_target,y_target,out

    lable_values = 'Test #','Car Angle','Car Distance','Angle to target','Desired angle turn','x_b','y_b','x_r','y_r','x_target','y_target','Output'

    i = 0
    
  
    for i in range(len(data_values)):
        i+=1
        #write lables on row 1 of xlsx file
        sheet.cell(row=1,column=i).value = lable_values[i-1]
        #write data on row 2 of xlsx file
        sheet.cell(row=y,column=i).value = data_values[i-1]
    

    


    #save the file and tell the terminal if it saved

    y+=1
wb.save(r'C:\Users\awind\Desktop\test1.xlsx')
print("Data saved to Excel file.")

wb.close()



#driver = webdriver.Chrome()
#driver.get("https://test4python.sarahah.com/")
#driver.find_element_by_tag_name("textarea").send_keys("This is the input text which is inputed.")







#C:\Users\awind\OneDrive - MNSCU\friend.xlsx
#C:\Users\awind\Desktop\Test_1.xlsx




















































#test = 0
#test_vals = [0,0,0,0,0,0]
#for test in range(0,1):
#    user_input = float(input("Enter data to insert into Excel (comma-separated): "))
#    test_vals[test] = user_input







#write lables on row 1 of xlsx file
#sheet.cell(row=1,column=1).value = 'Car Angle'
#sheet.cell(row=1,column=2).value = 'Car Distance'
#sheet.cell(row=1,column=3).value = 'Angle to target'
#sheet.cell(row=1,column=4).value = 'Desired angle turn'
#sheet.cell(row=1,column=5).value = 'x_b'
#sheet.cell(row=1,column=6).value = 'y_b'
#sheet.cell(row=1,column=7).value = 'x_r'
#sheet.cell(row=1,column=8).value = 'x_r'
#sheet.cell(row=1,column=9).value = 'Output'


#write data on row 2 of xlsx file
#sheet.cell(row=2,column=1).value = data_values[0]
#sheet.cell(row=2,column=2).value = data_values[1]
#sheet.cell(row=2,column=3).value = data_values[2]
#sheet.cell(row=2,column=4).value = data_values[3]
#sheet.cell(row=2,column=5).value = data_values[4]
#sheet.cell(row=2,column=6).value = data_values[5]
#sheet.cell(row=2,column=7).value = data_values[6]
#sheet.cell(row=2,column=8).value = data_values[7]    
#sheet.cell(row=2,column=9).value = data_values[8] 
















#from ultralytics import YOLO
#from ultralytics.yolo.v8.detect.predict import DetectionPredictor
#import cv2
#loading model
#model = YOLO('yolov8n.pt')
#perform tracking with the model
#results = model.track(source = 0, show=True, tracker = "bytetrack.yaml") #tracking with default tracker